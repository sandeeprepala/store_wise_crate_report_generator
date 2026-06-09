import openpyxl
from pathlib import Path
p = Path(r'c:\Users\sande\Desktop\Projects\daddy\Crates holding details MM Wise.xlsx')
wb = openpyxl.load_workbook(p, data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\nSheet:', name)
    for r in range(1, 40):
        row = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        if any(cell is not None for cell in row):
            print(r, row)
